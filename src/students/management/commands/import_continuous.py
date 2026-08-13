from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone
from decimal import Decimal
from datetime import datetime, timedelta
from students.models import Program, Cohort, Student, Course, Enrollment, Payment, ContinuousImportLog
import openpyxl
import logging

logger = logging.getLogger('imports')


class Command(BaseCommand):
    help = 'Import continuous data from Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Path to Excel file')
        parser.add_argument('--dry-run', action='store_true', help='Preview without saving')
        parser.add_argument('--update-existing', action='store_true', help='Update existing records')
        parser.add_argument('--import-name', type=str, required=True, help='Batch identifier')

    def handle(self, *args, **options):
        file_path = options['file_path']
        dry_run = options['dry_run']
        update_existing = options['update_existing']
        import_name = options['import_name']

        self.stdout.write(f"Starting import: {import_name}")
        if dry_run:
            self.stdout.write("DRY RUN MODE - No changes will be saved")

        stats = {
            'programs': {'created': 0, 'updated': 0, 'skipped': 0},
            'cohorts': {'created': 0, 'updated': 0, 'skipped': 0},
            'students': {'created': 0, 'updated': 0, 'skipped': 0},
            'courses': {'created': 0, 'updated': 0, 'skipped': 0},
            'enrollments': {'created': 0, 'updated': 0, 'skipped': 0},
            'payments': {'created': 0, 'updated': 0, 'skipped': 0},
        }

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            self.stderr.write(f"Error opening file: {e}")
            return

        with transaction.atomic():
            if 'Programs' in wb.sheetnames:
                stats['programs'] = self.import_programs(wb['Programs'], dry_run, update_existing)
            if 'Cohorts' in wb.sheetnames:
                stats['cohorts'] = self.import_cohorts(wb['Cohorts'], dry_run, update_existing)
            if 'Students' in wb.sheetnames:
                stats['students'] = self.import_students(wb['Students'], dry_run, update_existing)
            if 'Courses' in wb.sheetnames:
                stats['courses'] = self.import_courses(wb['Courses'], dry_run, update_existing)
            if 'Enrollments' in wb.sheetnames:
                stats['enrollments'] = self.import_enrollments(wb['Enrollments'], dry_run, update_existing)
            if 'Payments' in wb.sheetnames:
                stats['payments'] = self.import_payments(wb['Payments'], dry_run, update_existing)

            if dry_run:
                transaction.set_rollback(True)

            log = ContinuousImportLog.objects.create(
                import_name=import_name,
                file_path=file_path,
                stats=stats,
                imported_by=self.get_user(),
                dry_run=dry_run,
            )

        self.stdout.write(self.style.SUCCESS(f"Import completed: {stats}"))
        logger.info(f"Import {import_name} completed with stats: {stats}")

    def import_programs(self, ws, dry_run, update_existing):
        stats = {'created': 0, 'updated': 0, 'skipped': 0}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            code, name, duration, total_fees, allows_installments = row[:5]
            try:
                duration = int(duration)
                total_fees = Decimal(str(total_fees))
                allows_installments = str(allows_installments).lower() in ['true', '1', 'yes']
            except (ValueError, TypeError):
                stats['skipped'] += 1
                continue
            if dry_run:
                stats['created'] += 1
                continue
            obj, created = Program.objects.update_or_create(
                code=code,
                defaults={
                    'name': name,
                    'duration_months': duration,
                    'total_fees': total_fees,
                    'allows_installments': allows_installments,
                }
            )
            if created:
                stats['created'] += 1
            else:
                stats['updated'] += 1
        return stats

    def import_cohorts(self, ws, dry_run, update_existing):
        stats = {'created': 0, 'updated': 0, 'skipped': 0}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name, program_code, start_date, expected_end_date, max_students = row[:5]
            try:
                program = Program.objects.get(code=program_code)
                start_date = datetime.strptime(str(start_date), '%Y-%m-%d').date()
                expected_end_date = datetime.strptime(str(expected_end_date), '%Y-%m-%d').date()
                max_students = int(max_students)
            except (Program.DoesNotExist, ValueError, TypeError):
                stats['skipped'] += 1
                continue
            if dry_run:
                stats['created'] += 1
                continue
            obj, created = Cohort.objects.update_or_create(
                name=name, program=program,
                defaults={
                    'start_date': start_date,
                    'expected_end_date': expected_end_date,
                    'max_students': max_students,
                }
            )
            if created:
                stats['created'] += 1
            else:
                stats['updated'] += 1
        return stats

    def import_students(self, ws, dry_run, update_existing):
        stats = {'created': 0, 'updated': 0, 'skipped': 0}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            student_id, first_name, last_name, email, phone, program_code, cohort_name, enrollment_date, status = row[:9]
            try:
                program = Program.objects.get(code=program_code)
                cohort = Cohort.objects.get(name=cohort_name, program=program)
                enrollment_date = datetime.strptime(str(enrollment_date), '%Y-%m-%d').date()
                expected_graduation_date = enrollment_date + timedelta(days=program.duration_months * 30)
            except (Program.DoesNotExist, Cohort.DoesNotExist, ValueError, TypeError):
                stats['skipped'] += 1
                continue
            if dry_run:
                stats['created'] += 1
                continue
            obj, created = Student.objects.update_or_create(
                student_id=student_id,
                defaults={
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email or '',
                    'phone': phone or '',
                    'program': program,
                    'cohort': cohort,
                    'enrollment_date': enrollment_date,
                    'expected_graduation_date': expected_graduation_date,
                    'status': status or 'ENROLLED',
                }
            )
            if created:
                stats['created'] += 1
            else:
                stats['updated'] += 1
        return stats

    def import_courses(self, ws, dry_run, update_existing):
        stats = {'created': 0, 'updated': 0, 'skipped': 0}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            course_code, title, credits, program_codes = row[:4]
            try:
                credits = int(credits)
            except (ValueError, TypeError):
                stats['skipped'] += 1
                continue
            if dry_run:
                stats['created'] += 1
                continue
            course, created = Course.objects.update_or_create(
                course_code=course_code,
                defaults={'title': title, 'credits': credits}
            )
            if program_codes:
                codes = [c.strip() for c in str(program_codes).split(',')]
                programs = Program.objects.filter(code__in=codes)
                course.programs.set(programs)
            if created:
                stats['created'] += 1
            else:
                stats['updated'] += 1
        return stats

    def import_enrollments(self, ws, dry_run, update_existing):
        stats = {'created': 0, 'updated': 0, 'skipped': 0}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            student_id, course_code, status, letter_grade, completion_date = row[:5]
            try:
                student = Student.objects.get(student_id=student_id)
                course = Course.objects.get(course_code=course_code)
                enrollment_date = student.enrollment_date
            except (Student.DoesNotExist, Course.DoesNotExist):
                stats['skipped'] += 1
                continue
            if dry_run:
                stats['created'] += 1
                continue
            obj, created = Enrollment.objects.update_or_create(
                student=student, course=course,
                defaults={
                    'enrollment_date': enrollment_date,
                    'status': status or 'ENROLLED',
                    'letter_grade': letter_grade or '',
                }
            )
            if created:
                stats['created'] += 1
            else:
                stats['updated'] += 1
        return stats

    def import_payments(self, ws, dry_run, update_existing):
        stats = {'created': 0, 'updated': 0, 'skipped': 0}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            student_id, program_code, amount, payment_date, verified, notes = row[:6]
            try:
                student = Student.objects.get(student_id=student_id)
                program = Program.objects.get(code=program_code)
                amount = Decimal(str(amount))
                payment_date = datetime.strptime(str(payment_date), '%Y-%m-%d').date()
                status = 'VERIFIED' if str(verified).lower() in ['true', '1', 'yes'] else 'PENDING'
            except (Student.DoesNotExist, Program.DoesNotExist, ValueError, TypeError):
                stats['skipped'] += 1
                continue
            if dry_run:
                stats['created'] += 1
                continue
            obj, created = Payment.objects.update_or_create(
                student=student, program=program, payment_date=payment_date, amount=amount,
                defaults={'status': status, 'notes': notes or ''}
            )
            if created:
                stats['created'] += 1
            else:
                stats['updated'] += 1
        return stats

    def get_user(self):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        return User.objects.filter(user_type='ADMIN').first()
